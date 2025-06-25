from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, date
import re
import os

# --- Configuración global ---
NOMBRE_ARCHIVO_EXCEL = "Listado_Empresas_ARL_Automatizado.xlsx"
ENCABEZADOS = [
    "ORG_JURIDICA", "FECHA_DE_MATRICULA", "RAZON_SOCIAL", "TIPO_DE_IDENTIFICACION",
    "NUMERO_DE_NIT", "CIIU", "INGRESOS", "DEPARTAMENTO", "MUNICIPIO",
    "DIRECCION", "CORREO", "TELEFONO", "PAGINA_WEB",
    "REPRESENTANTE_LEGAL", "TIPO_DE_RIESGO_ARL"
]

# --- Funciones de Utilidad ---

def aplicar_estilos_encabezados(ws):
    """Aplica estilos a los encabezados y ajusta el ancho de las columnas."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    for col_idx, cell in enumerate(ws[1]):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Calcular el ancho máximo basado en el encabezado y los datos existentes
        max_length = len(ENCABEZADOS[col_idx])
        for row_idx in range(2, ws.max_row + 1): # Revisa desde la fila 2 (datos)
            cell_value = ws.cell(row=row_idx, column=col_idx + 1).value
            if cell_value:
                # Asegura que las fechas se conviertan a cadena para medir la longitud
                if isinstance(cell_value, (datetime, date)):
                    cell_value = cell_value.strftime("%d/%m/%Y")
                max_length = max(max_length, len(str(cell_value)))
        ws.column_dimensions[cell.column_letter].width = max_length + 2 # +2 para un pequeño margen

def validar_campo(nombre_campo, valor_ingresado, requerido=True):
    """Valida un valor de campo específico y devuelve el valor convertido o un error."""
    if requerido and not valor_ingresado and valor_ingresado != "": # Asegurarse que '' es válido para no requerido
        return None, f"El campo '{nombre_campo}' es obligatorio."

    # Permitir vacío para campos no requeridos si el valor ingresado es una cadena vacía
    if not valor_ingresado and not requerido and valor_ingresado == "":
        return valor_ingresado, None

    if nombre_campo == "FECHA_DE_MATRICULA":
        try:
            return datetime.strptime(valor_ingresado, "%d/%m/%Y").date(), None
        except ValueError:
            return None, "Formato de fecha incorrecto. Por favor, use DD/MM/AAAA."
    elif nombre_campo == "INGRESOS":
        try:
            return float(valor_ingresado), None
        except ValueError:
            return None, "Valor de ingresos inválido. Por favor, ingrese solo números."
    elif nombre_campo == "CORREO":
        if valor_ingresado and not re.match(r"[^@]+@[^@]+\.[^@]+", valor_ingresado):
            return None, "Formato de correo electrónico inválido."
        return valor_ingresado, None
    elif nombre_campo == "TELEFONO":
        if valor_ingresado and not re.fullmatch(r'\d{7}|\d{10}', valor_ingresado):
            return None, "Formato de teléfono inválido (solo números, 7 o 10 dígitos)."
        return valor_ingresado, None
    elif nombre_campo == "TIPO_DE_IDENTIFICACION":
        if valor_ingresado.upper() not in ["NIT", "NO NIT"]:
            return None, "Tipo de identificación inválido. Debe ser 'NIT' o 'NO NIT'."
        return valor_ingresado.upper(), None
    elif nombre_campo == "NUMERO_DE_NIT":
        # Esta validación se usa si el campo se ingresa, la lógica de salto está en obtener_datos_empresa_manual
        # Permite formato "900123456-7" o solo números
        if valor_ingresado.upper() == "N/A" or not valor_ingresado: # "N/A" o vacío siempre son válidos aquí si no es NIT
             return valor_ingresado.upper() if valor_ingresado else "", None
        
        # Si no es N/A ni vacío, validar formato NIT
        if not re.fullmatch(r'^\d{9}-\d$', valor_ingresado) and not re.fullmatch(r'^\d+$', valor_ingresado):
            return None, "Formato de NIT inválido. Use 'XXXXXXXXX-X' o solo números."
        return valor_ingresado, None
    elif nombre_campo == "CIIU":
        # CIIU: entre 4 y 5 dígitos numéricos
        if not re.fullmatch(r'^\d{4,5}$', valor_ingresado): # MODIFICADO: Acepta 4 o 5 dígitos
            return None, "Formato de CIIU inválido. Debe ser un número de 4 o 5 dígitos."
        return int(valor_ingresado), None
    
    return valor_ingresado, None

# --- Funciones de Gestión de Empresas ---

def obtener_datos_empresa_manual(modo="agregar", datos_actuales=None):
    """
    Solicita al usuario los datos de una empresa campo por campo.
    Si modo="actualizar", muestra los datos actuales como referencia.
    """
    datos = {}
    print("\n" + "~"*60)
    print(f"--- {('ACTUALIZANDO' if modo == 'actualizar' else 'INGRESANDO NUEVOS')} DATOS DE EMPRESA ---".center(60))
    print("~"*60 + "\n")

    campos_info = {
        "ORG_JURIDICA": ("Tipo de organización (Ej: Persona Natural, Persona Juridica, Est. Ag. Suc)", True),
        "FECHA_DE_MATRICULA": ("Fecha de Matrícula (Formato: DD/MM/AAAA, Ej: 24/06/2025)", True),
        "RAZON_SOCIAL": ("Nombre de la empresa", True),
        "TIPO_DE_IDENTIFICACION": ("Tipo de identificación (Ej: NIT, NO NIT)", True),
        "NUMERO_DE_NIT": ("Número de NIT/Identificación (Ej: 900123456-7 o N/A si es NO NIT)", True),
        "CIIU": ("CIIU (Actividad económica, solo 4 o 5 dígitos numéricos, Ej: 62090)", True), # Mensaje actualizado
        "INGRESOS": ("Ingresos (Solo números, Ej: 125000000)", True),
        "DEPARTAMENTO": ("Departamento (Ej: Bolivar, Cundinamarca)", True),
        "MUNICIPIO": ("Municipio (Ej: Cartagena, Bogota)", True),
        "DIRECCION": ("Dirección completa", True),
        "CORREO": ("Correo (Ej: info@empresa.com)", False),
        "TELEFONO": ("Teléfono (Solo números, 7 o 10 dígitos, Ej: 3101234567)", False),
        "PAGINA_WEB": ("Página Web (Opcional, Ej: www.empresa.com)", False),
        "REPRESENTANTE_LEGAL": ("Representante Legal", True),
        "TIPO_DE_RIESGO_ARL": ("Tipo de Riesgo ARL (Ej: Riesgo I, Riesgo V, Sin Riesgo)", True)
    }
    
    nombres_campos_ordenados = ENCABEZADOS

    # Variables para validación cruzada y control de flujo
    tipo_identificacion_ingresado = None

    for i, nombre_campo in enumerate(nombres_campos_ordenados):
        mensaje, requerido = campos_info.get(nombre_campo, ("", True))

        # Manejo especial para NUMERO_DE_NIT
        if nombre_campo == "NUMERO_DE_NIT":
            if tipo_identificacion_ingresado == "NO NIT":
                # Si es NO NIT, no preguntar por el NIT, directamente asignar "N/A"
                print(f"  {i+1}. {nombre_campo}: Asignando 'N/A' (Tipo de Identificación es 'NO NIT')")
                datos[nombre_campo] = "N/A"
                continue # Saltar a la siguiente iteración del bucle
            elif modo == "actualizar":
                # Si es NIT y estamos actualizando, necesitamos el tipo de identificación actual
                idx_tipo_id_actual = ENCABEZADOS.index("TIPO_DE_IDENTIFICACION")
                if datos_actuales and datos_actuales[idx_tipo_id_actual] == "NO NIT":
                    print(f"  {i+1}. {nombre_campo}: Asignando 'N/A' (Tipo de Identificación actual es 'NO NIT')")
                    datos[nombre_campo] = "N/A"
                    continue # Saltar a la siguiente iteración

        valor_actual = datos_actuales[i] if datos_actuales and i < len(datos_actuales) else ""
        if isinstance(valor_actual, (datetime, date)):
            valor_actual = valor_actual.strftime("%d/%m/%Y")
        
        prompt = f"  {i+1}. {nombre_campo} ({mensaje})"
        if modo == "actualizar":
            prompt += f" [Actual: {valor_actual}]"
        prompt += ": "

        while True:
            valor_ingresado = input(prompt).strip()
            if valor_ingresado.lower() == 'fin':
                return None

            if modo == "actualizar" and not valor_ingresado:
                # Mantener el valor actual y capturar el tipo_identificacion si aplica
                datos[nombre_campo] = datos_actuales[i]
                if nombre_campo == "TIPO_DE_IDENTIFICACION":
                    tipo_identificacion_ingresado = datos_actuales[i]
                break
            
            valor_validado, error = validar_campo(nombre_campo, valor_ingresado, requerido)
            
            if nombre_campo == "TIPO_DE_IDENTIFICACION":
                tipo_identificacion_ingresado = valor_validado
                
            # Validacion cruzada de NUMERO_DE_NIT DESPUÉS de su propia validación de formato
            if nombre_campo == "NUMERO_DE_NIT":
                # Si TIPO_DE_IDENTIFICACION es NIT, NUMERO_DE_NIT no puede ser N/A o vacío
                if tipo_identificacion_ingresado == "NIT" and (not valor_validado or str(valor_validado).upper() == "N/A"):
                    error = "Si el Tipo de Identificación es 'NIT', el Número de NIT no puede ser 'N/A' o vacío."
                    valor_validado = None
                elif tipo_identificacion_ingresado == "NO NIT" and valor_validado not in ["N/A", ""]:
                    # Esto solo ocurriría si el usuario explícitamente puso algo diferente a N/A/vacío
                    # Aunque la lógica de salto ya debería evitar que pregunte
                    error = "Si el Tipo de Identificación es 'NO NIT', el Número de NIT debe ser 'N/A' o vacío."
                    valor_validado = None

            if error:
                print(f"    * {error} Intente de nuevo.")
                continue
            else:
                datos[nombre_campo] = valor_validado
                break
    
    return datos

def agregar_empresa(ws):
    """Función para agregar una nueva empresa manualmente."""
    nueva_empresa_data = obtener_datos_empresa_manual(modo="agregar")
    if nueva_empresa_data is None:
        print("Operación de adición cancelada.")
        return False

    fila_a_agregar = [nueva_empresa_data[h] for h in ENCABEZADOS]
    ws.append(fila_a_agregar)
    print(f"\n¡Empresa '{nueva_empresa_data['RAZON_SOCIAL']}' agregada con éxito!")
    return True

def actualizar_empresa_interactivo(ws):
    """
    Permite actualizar una empresa existente mostrando un menú de campos.
    """
    print("\n" + "="*60)
    print("--- ACTUALIZAR EMPRESA EXISTENTE ---".center(60))
    print("="*60 + "\n")

    nombre_empresa_busqueda = input("Ingrese la RAZON SOCIAL de la empresa a actualizar (o 'fin' para cancelar): ").strip()
    if not nombre_empresa_busqueda or nombre_empresa_busqueda.lower() == 'fin':
        print("Operación de actualización cancelada.")
        return False

    fila_encontrada = -1
    datos_empresa_actuales = []
    idx_razon_social = ENCABEZADOS.index("RAZON_SOCIAL")
    idx_tipo_identificacion = ENCABEZADOS.index("TIPO_DE_IDENTIFICACION")
    idx_numero_nit = ENCABEZADOS.index("NUMERO_DE_NIT")


    for i, row in enumerate(ws.iter_rows(min_row=2)):
        row_values = [cell.value for cell in row]
        if str(row_values[idx_razon_social]).lower() == nombre_empresa_busqueda.lower():
            fila_encontrada = i + 2
            datos_empresa_actuales = row_values
            break

    if fila_encontrada == -1:
        print(f"\nNo se encontró ninguna empresa con la Razón Social: '{nombre_empresa_busqueda}'.")
        return False

    print(f"\nEmpresa encontrada en la fila {fila_encontrada}:")
    for j, header in enumerate(ENCABEZADOS):
        valor = datos_empresa_actuales[j]
        if isinstance(valor, (datetime, date)):
            valor = valor.strftime("%d/%m/%Y")
        print(f"  {header}: {valor}")
    print("-" * 60)

    # --- Menú de Campos a Actualizar ---
    while True:
        print("\nSeleccione el campo a actualizar (o '0' para terminar de actualizar esta empresa):")
        for i, header in enumerate(ENCABEZADOS):
            print(f"  {i+1}. {header}")
        
        opcion_campo = input("Ingrese el número del campo: ").strip()

        if opcion_campo == '0':
            print(f"\nTerminando actualización para '{nombre_empresa_busqueda}'.")
            break

        try:
            indice_campo = int(opcion_campo) - 1
            if not (0 <= indice_campo < len(ENCABEZADOS)):
                raise ValueError
        except ValueError:
            print("Opción inválida. Por favor, ingrese un número del 1 al", len(ENCABEZADOS), "o '0'.")
            continue

        campo_a_actualizar = ENCABEZADOS[indice_campo]
        
        # Obtener el valor actual del tipo de identificación para validación cruzada
        # Si se va a actualizar TIPO_DE_IDENTIFICACION, usamos el nuevo valor ingresado
        # Si no, usamos el valor actual de la empresa
        tipo_identificacion_para_validacion = datos_empresa_actuales[idx_tipo_identificacion]
        
        # Si el campo a actualizar es TIPO_DE_IDENTIFICACION, el nuevo valor ingresado será el 'tipo_identificacion_para_validacion'
        # Esto se manejará en la lógica de entrada, pero lo necesitamos aquí para la validación del NIT

        print(f"\n-> Actualizando: '{campo_a_actualizar}' [Actual: {datos_empresa_actuales[indice_campo]}]")
        nuevo_valor_str = input(f"  Ingrese el nuevo valor para '{campo_a_actualizar}': ").strip()
        
        es_requerido = campo_a_actualizar not in ["CORREO", "TELEFONO", "PAGINA_WEB"]
        
        valor_validado, error = validar_campo(campo_a_actualizar, nuevo_valor_str, es_requerido)

        # Lógica de validación cruzada para actualizar
        if campo_a_actualizar == "TIPO_DE_IDENTIFICACION":
            tipo_identificacion_para_validacion = valor_validado # Usar el nuevo valor del tipo de identificación
            if tipo_identificacion_para_validacion == "NO NIT":
                # Si cambia a NO NIT, forzar NIT a N/A y actualizar en excel
                print("ADVERTENCIA: Si el tipo de identificación es 'NO NIT', el Número de NIT se establecerá a 'N/A'.")
                ws.cell(row=fila_encontrada, column=idx_numero_nit + 1, value="N/A")
                datos_empresa_actuales[idx_numero_nit] = "N/A" # Actualizar en la lista temporal
            elif tipo_identificacion_para_validacion == "NIT":
                # Si cambia a NIT, el NIT no puede ser N/A o vacío
                if not datos_empresa_actuales[idx_numero_nit] or str(datos_empresa_actuales[idx_numero_nit]).upper() == "N/A":
                    error = "Si el Tipo de Identificación es 'NIT', el Número de NIT no puede ser vacío o 'N/A'. Por favor, actualice el NIT."
                    valor_validado = None # Invalidar el cambio de tipo si el NIT no es válido

        elif campo_a_actualizar == "NUMERO_DE_NIT":
            if tipo_identificacion_para_validacion == "NO NIT":
                if nuevo_valor_str.upper() != "N/A" and nuevo_valor_str != "":
                    error = "Si el Tipo de Identificación es 'NO NIT', el Número de NIT debe ser 'N/A' o vacío."
                    valor_validado = None
                else:
                    valor_validado = "N/A" # Forzar a "N/A"
                    error = None
            elif tipo_identificacion_para_validacion == "NIT" and (not nuevo_valor_str or nuevo_valor_str.upper() == "N/A"):
                error = "Si el Tipo de Identificación es 'NIT', el Número de NIT no puede ser 'N/A' o vacío."
                valor_validado = None

        if error:
            print(f"    * Error al actualizar '{campo_a_actualizar}': {error}")
        else:
            datos_empresa_actuales[indice_campo] = valor_validado
            ws.cell(row=fila_encontrada, column=indice_campo + 1, value=valor_validado)
            print(f"    '{campo_a_actualizar}' actualizado exitosamente.")
    
    print(f"\n¡Empresa '{datos_empresa_actuales[idx_razon_social]}' en la fila {fila_encontrada} actualizada exitosamente!")
    return True

def cargar_multiples_empresas(ws):
    """
    Permite al usuario pegar múltiples líneas de empresas desde la consola.
    """
    print("\n" + "="*60)
    print("--- CARGA MASIVA DE EMPRESAS ---".center(60))
    print("="*60 + "\n")
    print("Por favor, pegue las líneas de las empresas en el siguiente formato:")
    print(" | ".join(ENCABEZADOS))
    print("Cada empresa en una nueva línea. Cuando termine, escriba 'FIN_CARGA' en una línea separada.\n")

    empresas_procesadas = 0
    errores_en_carga = 0
    
    while True:
        linea = input("Pegue línea de empresa o 'FIN_CARGA': ").strip()
        
        if linea.lower() == 'fin_carga':
            break
        if not linea: # Ignorar líneas vacías accidentales
            continue

        partes = linea.split('|')

        if len(partes) != len(ENCABEZADOS):
            print(f"ERROR: Línea inválida (campos incorrectos: {len(partes)} vs {len(ENCABEZADOS)} esperados): {linea}")
            errores_en_carga += 1
            continue
        
        datos_para_fila = {}
        fila_valida = True
        
        # Pre-procesar TIPO_DE_IDENTIFICACION para la validación cruzada del NIT
        tipo_identificacion_temp = None
        try:
            idx_tipo_id = ENCABEZADOS.index("TIPO_DE_IDENTIFICACION")
            valor_tipo_id_str = partes[idx_tipo_id].strip()
            tipo_identificacion_temp, error_tipo_id = validar_campo("TIPO_DE_IDENTIFICACION", valor_tipo_id_str, True)
            if error_tipo_id:
                print(f"ERROR: En línea '{linea}' -> Campo 'TIPO_DE_IDENTIFICACION': {error_tipo_id}")
                fila_valida = False
            else:
                datos_para_fila["TIPO_DE_IDENTIFICACION"] = tipo_identificacion_temp
        except ValueError:
            fila_valida = False
            print(f"ERROR: No se encontró el campo TIPO_DE_IDENTIFICACION en los encabezados.")

        if not fila_valida:
            errores_en_carga += 1
            continue

        for i, header in enumerate(ENCABEZADOS):
            valor_str = partes[i].strip()
            es_requerido = header not in ["CORREO", "TELEFONO", "PAGINA_WEB"]

            # Lógica para NO preguntar por NIT en carga masiva si TIPO_DE_IDENTIFICACION es NO NIT
            if header == "NUMERO_DE_NIT":
                if tipo_identificacion_temp == "NO NIT":
                    valor_validado = "N/A" # Asignar N/A directamente
                    error = None # No hay error de formato si forzamos N/A
                else: # Si es NIT, validar el número de NIT normalmente
                    valor_validado, error = validar_campo(header, valor_str, es_requerido)
                    if tipo_identificacion_temp == "NIT" and (not valor_validado or str(valor_validado).upper() == "N/A"):
                        error = "Si el Tipo de Identificación es 'NIT', el Número de NIT no puede ser 'N/A' o vacío."
                        valor_validado = None
            else:
                valor_validado, error = validar_campo(header, valor_str, es_requerido)

            if error:
                print(f"ERROR: En línea '{linea}' -> Campo '{header}': {error}")
                fila_valida = False
                break
            datos_para_fila[header] = valor_validado

        if fila_valida:
            try:
                fila_ordenada = [datos_para_fila[h] for h in ENCABEZADOS]
                ws.append(fila_ordenada)
                empresas_procesadas += 1
            except Exception as e:
                print(f"ERROR: No se pudo añadir la línea al Excel: {linea} - {e}")
                errores_en_carga += 1
        else:
            errores_en_carga += 1
            
    print("\n" + "="*60)
    print("--- RESUMEN DE LA CARGA MASIVA ---".center(60))
    print(f"Empresas añadidas exitosamente: {empresas_procesadas}".center(60))
    print(f"Errores encontrados: {errores_en_carga}".center(60))
    print("="*60 + "\n")
    return True

# --- Menú Principal ---

def iniciar_gestion_empresas():
    """Función principal que inicia el programa y muestra el menú inicial."""
    
    if os.path.exists(NOMBRE_ARCHIVO_EXCEL):
        print(f"\nCargando archivo existente: {NOMBRE_ARCHIVO_EXCEL}")
        try:
            wb = load_workbook(NOMBRE_ARCHIVO_EXCEL)
            ws = wb.active
            if [cell.value for cell in ws[1]] != ENCABEZADOS:
                print("--- ATENCIÓN ---".center(60))
                print("Advertencia: Los encabezados del archivo existente no coinciden con los esperados.")
                print("Esto podría causar problemas. Por favor, revise el archivo o considere iniciar uno nuevo.")
                print("----------------".center(60))
            aplicar_estilos_encabezados(ws)
        except Exception as e:
            print(f"Error al cargar el archivo existente: {e}")
            print("Creando un nuevo archivo en su lugar.")
            wb = Workbook()
            ws = wb.active
            ws.title = "Empresas_ARL"
            ws.append(ENCABEZADOS)
            aplicar_estilos_encabezados(ws)
    else:
        print(f"\nEl archivo '{NOMBRE_ARCHIVO_EXCEL}' no existe. Creando uno nuevo.")
        wb = Workbook()
        ws = wb.active
        ws.title = "Empresas_ARL"
        ws.append(ENCABEZADOS)
        aplicar_estilos_encabezados(ws)

    while True:
        print("\n" + "="*60)
        print("--- GESTIÓN DE EMPRESAS ---".center(60))
        print("="*60)
        print("1. Agregar nueva empresa (manual)")
        print("2. Actualizar empresa existente (por Razón Social)")
        print("3. Cargar múltiples empresas (desde consola)")
        print("4. Salir y Guardar")
        print("="*60)

        opcion = input("Seleccione una opción: ").strip()

        if opcion == '1':
            if agregar_empresa(ws):
                aplicar_estilos_encabezados(ws)
        elif opcion == '2':
            if actualizar_empresa_interactivo(ws):
                aplicar_estilos_encabezados(ws)
        elif opcion == '3':
            if cargar_multiples_empresas(ws):
                aplicar_estilos_encabezados(ws)
        elif opcion == '4':
            try:
                wb.save(NOMBRE_ARCHIVO_EXCEL)
                print(f"\n¡Cambios guardados en '{NOMBRE_ARCHIVO_EXCEL}' y saliendo del programa!")
            except Exception as e:
                print(f"Error al guardar el archivo: {e}. Asegúrese de que no esté abierto en Excel.")
            break
        else:
            print("Opción inválida. Por favor, intente de nuevo.")

if __name__ == "__main__":
    iniciar_gestion_empresas()